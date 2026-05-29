"""I1: Logo Tiger ERP connector — XML mode.

## Logo Tiger XML formatı

Logo'nun "Veri Aktarımı > XML" feature'u standart şu yapıda çıktı verir:

  <LOGOWORLD>
    <CARI>
      <CARI_KODU>CR001</CARI_KODU>
      <CARI_UNVAN>Acme Ticaret Ltd. Şti.</CARI_UNVAN>
      <CARI_VKN>1234567890</CARI_VKN>
      <CARI_VERGI_DAIRESI>Kadıköy</CARI_VERGI_DAIRESI>
      <CARI_ADRES>...</CARI_ADRES>
      <CARI_TELEFON>...</CARI_TELEFON>
      <CARI_EMAIL>...</CARI_EMAIL>
      <CARI_IBAN>...</CARI_IBAN>
      <CARI_BAKIYE>15000.50</CARI_BAKIYE>
      <CARI_DOVIZ>TRY</CARI_DOVIZ>
      <CARI_TIPI>1</CARI_TIPI>     <!-- 1=Müşteri 2=Tedarikçi 3=Her ikisi -->
    </CARI>
    <FATURA>
      <FATURA_NO>F2026000001</FATURA_NO>
      <FATURA_CARI_KODU>CR001</FATURA_CARI_KODU>
      <FATURA_TARIHI>2026-05-15</FATURA_TARIHI>
      <FATURA_VADE>2026-06-15</FATURA_VADE>
      <FATURA_NET>10000.00</FATURA_NET>
      <FATURA_KDV>1800.00</FATURA_KDV>
      <FATURA_BRUT>11800.00</FATURA_BRUT>
      <FATURA_DOVIZ>TRY</FATURA_DOVIZ>
      <FATURA_TIPI>1</FATURA_TIPI>  <!-- 1=Satış 2=Alış -->
      <FATURA_ACIKLAMA>...</FATURA_ACIKLAMA>
    </FATURA>
  </LOGOWORLD>

Bu modül stdlib xml.etree ile parse eder — zero external dep.

## Excel mode (placeholder)

openpyxl gerekli; lazy import. Sheet'ler: "Cariler", "Faturalar".
Header row Logo XML'in element adlarına denk gelir (Tiger default).

## Field tolerance

Logo versiyonları arasında küçük farklar olabilir (örn. CARI_TIPI vs
CARI_TYPE). Alternatif key listesi ile graceful degradation.
"""
from __future__ import annotations

import hashlib
import re
from typing import Any

# xml.etree.ElementTree expat'a bağımlı; bazı Python kurulumlarında
# (Homebrew Python örneği) expat yüklenemez. Pure-Python fallback ile
# bu durumu nazikçe halleder.
try:
    from xml.etree import ElementTree as ET  # type: ignore[import-not-found]
    _ETREE_AVAILABLE = True
except ImportError:
    _ETREE_AVAILABLE = False
    ET = None  # type: ignore[assignment]


class _MiniXMLElement:
    """ElementTree.Element ile uyumlu küçük arayüz.

    Logo'nun düz hiyerarşili XML formatı için yeterli; namespace yok,
    attribute yok, CDATA yok. find()/findall() temel davranışları.
    """

    __slots__ = ("tag", "text", "children")

    def __init__(self, tag: str, text: str | None = None) -> None:
        self.tag = tag
        self.text: str | None = text
        self.children: list[_MiniXMLElement] = []

    def find(self, tag: str) -> "_MiniXMLElement | None":
        for c in self.children:
            if c.tag == tag:
                return c
        return None

    def findall(self, path: str) -> list["_MiniXMLElement"]:
        """`.//TAG` veya `TAG` desteği — Logo için yeterli."""
        if path.startswith(".//"):
            tag = path[3:]
            return self._descendants(tag)
        return [c for c in self.children if c.tag == path]

    def _descendants(self, tag: str) -> list["_MiniXMLElement"]:
        out: list[_MiniXMLElement] = []
        for c in self.children:
            if c.tag == tag:
                out.append(c)
            out.extend(c._descendants(tag))
        return out

    def __iter__(self):
        return iter(self.children)

    def __len__(self) -> int:
        return len(self.children)


def _parse_xml_pure(data: bytes) -> _MiniXMLElement:
    """Logo'nun düz XML formatı için minimal pure-Python parser.

    Destekler:
      * <TAG>value</TAG>
      * <TAG><CHILD>...</CHILD></TAG> (nested)
      * <?xml ...?> header (atlanır)
      * <!-- yorum --> (atlanır)
      * <TAG/> self-closing (boş element)
    Desteklemez: attributes, namespaces, CDATA.

    Logo XML'i bu kısıtların hepsine uyar.
    """
    text = data.decode("utf-8", errors="replace")
    # Header + yorum strip
    text = re.sub(r"<\?xml.*?\?>", "", text, flags=re.DOTALL)
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)

    pos = 0
    n = len(text)

    def skip_ws() -> None:
        nonlocal pos
        while pos < n and text[pos] in " \t\r\n":
            pos += 1

    def parse_element() -> _MiniXMLElement | None:
        nonlocal pos
        skip_ws()
        if pos >= n or text[pos] != "<":
            return None
        # Open tag
        end_open = text.find(">", pos)
        if end_open < 0:
            raise ValueError("Beklenen kapanış '>' bulunamadı")
        tag_content = text[pos + 1:end_open].strip()
        pos = end_open + 1
        # Self-closing <TAG/>
        if tag_content.endswith("/"):
            return _MiniXMLElement(tag_content[:-1].strip(), text=None)
        tag = tag_content.split()[0]  # attribute varsa ilk word
        elem = _MiniXMLElement(tag)
        # Inner: text + children
        close_marker = f"</{tag}>"
        text_parts: list[str] = []
        closed = False
        while pos < n:
            skip_ws()
            if pos >= n:
                break
            if text.startswith(close_marker, pos):
                pos += len(close_marker)
                closed = True
                break
            if text[pos] == "<":
                # Yorum?
                if text.startswith("<!--", pos):
                    end_comment = text.find("-->", pos)
                    if end_comment < 0:
                        raise ValueError("Yorum kapanışı yok")
                    pos = end_comment + 3
                    continue
                child = parse_element()
                if child is not None:
                    elem.children.append(child)
            else:
                # Text content — sıradaki '<' kadar oku
                next_lt = text.find("<", pos)
                if next_lt < 0:
                    next_lt = n
                text_parts.append(text[pos:next_lt])
                pos = next_lt
        if not closed:
            raise ValueError(f"Beklenen </{tag}> bulunamadı")
        if not elem.children:
            joined = "".join(text_parts).strip()
            elem.text = joined or None
        return elem

    root = parse_element()
    if root is None:
        raise ValueError("Root element bulunamadı")
    return root

from app.connectors.base import (
    BaseConnector,
    ConnectorMode,
    ParsedCustomer,
    ParsedInvoice,
    ParseError,
    ParseResult,
)


# Alternatif tag isimleri — Logo versiyon farklarına karşı
CUSTOMER_KEYS: dict[str, tuple[str, ...]] = {
    "code":       ("CARI_KODU", "CARI_CODE", "CODE", "KOD"),
    "name":       ("CARI_UNVAN", "CARI_AD", "UNVAN", "AD", "NAME"),
    "tax_number": ("CARI_VKN", "CARI_TCKN", "VKN", "TCKN", "TAX_NO"),
    "tax_office": ("CARI_VERGI_DAIRESI", "VERGI_DAIRESI", "TAX_OFFICE"),
    "address":    ("CARI_ADRES", "ADRES", "ADDRESS"),
    "phone":      ("CARI_TELEFON", "TELEFON", "PHONE"),
    "email":      ("CARI_EMAIL", "EMAIL"),
    "iban":       ("CARI_IBAN", "IBAN"),
    "balance":    ("CARI_BAKIYE", "BAKIYE", "BALANCE"),
    "currency":   ("CARI_DOVIZ", "DOVIZ", "CURRENCY"),
    "type":       ("CARI_TIPI", "CARI_TYPE", "TIP", "TYPE"),
}

INVOICE_KEYS: dict[str, tuple[str, ...]] = {
    "no":          ("FATURA_NO", "FATURA_NUMARASI", "NO", "NUMBER"),
    "customer":    ("FATURA_CARI_KODU", "CARI_KODU"),
    "issue_date":  ("FATURA_TARIHI", "TARIH", "DATE"),
    "due_date":    ("FATURA_VADE", "VADE", "DUE_DATE"),
    "net":         ("FATURA_NET", "NET", "TUTAR"),
    "tax":         ("FATURA_KDV", "KDV", "VAT", "TAX"),
    "gross":       ("FATURA_BRUT", "BRUT", "GROSS"),
    "currency":    ("FATURA_DOVIZ", "DOVIZ", "CURRENCY"),
    "type":        ("FATURA_TIPI", "FATURA_TYPE", "TIP", "TYPE"),
    "description": ("FATURA_ACIKLAMA", "ACIKLAMA", "DESCRIPTION"),
}


class LogoTigerConnector(BaseConnector):
    """Logo Tiger ERP connector — XML + Excel modes."""

    connector_type = "logo_tiger"
    supported_modes = (ConnectorMode.XML, ConnectorMode.EXCEL)

    def parse(
        self,
        *,
        data: bytes,
        mode: ConnectorMode,
        filename: str | None = None,
    ) -> ParseResult:
        if mode == ConnectorMode.XML:
            return self._parse_xml(data)
        if mode == ConnectorMode.EXCEL:
            return self._parse_excel(data)
        if mode == ConnectorMode.WEB_SERVICE:
            raise NotImplementedError(
                "Logo Web Service modu henüz aktif değil — resmi API erişimi gerekli."
            )
        raise ValueError(f"Desteklenmeyen mod: {mode}")

    # ── XML mode ───────────────────────────────────────────────────────

    def _parse_xml(self, data: bytes) -> ParseResult:
        result = ParseResult()
        root: Any = None
        # Önce stdlib (hızlı + tam uyumlu), expat yoksa pure-Python fallback
        if _ETREE_AVAILABLE:
            try:
                root = ET.fromstring(data)
            except Exception as exc:  # ET.ParseError veya expat yokluğu
                root = None
                fallback_error: Exception | None = exc
        else:
            fallback_error = None
        if root is None:
            try:
                root = _parse_xml_pure(data)
            except (ValueError, Exception) as exc:
                result.errors.append(ParseError(
                    row_index=0,
                    record_type="root",
                    error_code="invalid_xml",
                    error_message=f"XML parse hatası: {exc}",
                ))
                return result

        result.source_info = {
            "root_tag": root.tag,
            "child_count": len(list(root)),
        }

        # Cariler — root altında veya <CARILER> wrapper içinde
        cari_nodes = self._collect_nodes(root, ("CARI", "CUSTOMER"))
        for idx, node in enumerate(cari_nodes):
            try:
                customer = self._parse_customer_node(node)
                if customer:
                    result.customers.append(customer)
            except (ValueError, TypeError) as exc:
                result.errors.append(ParseError(
                    row_index=idx,
                    record_type="customer",
                    error_code="invalid_customer",
                    error_message=str(exc),
                    raw_payload=_node_to_str(node)[:500],
                ))

        # Faturalar
        fatura_nodes = self._collect_nodes(root, ("FATURA", "INVOICE"))
        for idx, node in enumerate(fatura_nodes):
            try:
                invoice = self._parse_invoice_node(node)
                if invoice:
                    result.invoices.append(invoice)
            except (ValueError, TypeError) as exc:
                result.errors.append(ParseError(
                    row_index=idx,
                    record_type="invoice",
                    error_code="invalid_invoice",
                    error_message=str(exc),
                    raw_payload=_node_to_str(node)[:500],
                ))

        return result

    @staticmethod
    def _collect_nodes(
        root: Any, tags: tuple[str, ...]
    ) -> list[Any]:
        """Tag'leri root'tan ve direkt child wrapper'lardan topla."""
        nodes: list[ET.Element] = []
        for tag in tags:
            nodes.extend(root.findall(f".//{tag}"))
        # Deduplicate (aynı node birden fazla tag matching'inden gelmiş olabilir)
        seen_ids = set()
        unique: list[ET.Element] = []
        for n in nodes:
            nid = id(n)
            if nid not in seen_ids:
                seen_ids.add(nid)
                unique.append(n)
        return unique

    def _parse_customer_node(self, node: Any) -> ParsedCustomer | None:
        code = self._first_value(node, CUSTOMER_KEYS["code"])
        name = self._first_value(node, CUSTOMER_KEYS["name"])
        if not code or not name:
            raise ValueError("Cari kodu veya unvan eksik")

        tax_number = self._first_value(node, CUSTOMER_KEYS["tax_number"])
        tax_office = self._first_value(node, CUSTOMER_KEYS["tax_office"])
        address = self._first_value(node, CUSTOMER_KEYS["address"])
        phone = self._first_value(node, CUSTOMER_KEYS["phone"])
        email = self._first_value(node, CUSTOMER_KEYS["email"])
        iban = self._first_value(node, CUSTOMER_KEYS["iban"])
        balance = self._first_float(node, CUSTOMER_KEYS["balance"]) or 0.0
        currency = self._first_value(node, CUSTOMER_KEYS["currency"]) or "TRY"
        type_code = self._first_value(node, CUSTOMER_KEYS["type"]) or ""

        role = self._role_from_type(type_code)
        sig = _signature_hash(code, tax_number or "", "customer")

        return ParsedCustomer(
            source_code=code.strip(),
            name=name.strip(),
            tax_number=self._clean_tax_number(tax_number),
            tax_office=tax_office.strip() if tax_office else None,
            role=role,
            address=address.strip() if address else None,
            email=email.strip() if email else None,
            phone=phone.strip() if phone else None,
            iban=self._clean_iban(iban),
            balance=balance,
            currency=currency.strip().upper() if currency else "TRY",
            signature_hash=sig,
        )

    def _parse_invoice_node(self, node: Any) -> ParsedInvoice | None:
        no = self._first_value(node, INVOICE_KEYS["no"])
        cari = self._first_value(node, INVOICE_KEYS["customer"])
        issue = self._first_value(node, INVOICE_KEYS["issue_date"])
        if not no or not cari or not issue:
            raise ValueError("Fatura no / cari kodu / tarih eksik")

        issue_clean = self._normalize_date(issue)
        if not issue_clean:
            raise ValueError(f"Geçersiz tarih: {issue!r}")

        due = self._first_value(node, INVOICE_KEYS["due_date"])
        due_clean = self._normalize_date(due) if due else None

        net = self._first_float(node, INVOICE_KEYS["net"]) or 0.0
        tax = self._first_float(node, INVOICE_KEYS["tax"]) or 0.0
        gross = self._first_float(node, INVOICE_KEYS["gross"]) or (net + tax)

        currency = self._first_value(node, INVOICE_KEYS["currency"]) or "TRY"
        type_code = self._first_value(node, INVOICE_KEYS["type"]) or ""
        direction = self._direction_from_type(type_code)
        description = self._first_value(node, INVOICE_KEYS["description"])

        sig = _signature_hash(no, direction, cari)

        return ParsedInvoice(
            source_no=no.strip(),
            customer_source_code=cari.strip(),
            issue_date=issue_clean,
            due_date=due_clean,
            total_excl_tax=net,
            tax_amount=tax,
            total_incl_tax=gross,
            currency=currency.strip().upper(),
            direction=direction,
            description=description.strip() if description else None,
            signature_hash=sig,
        )

    @staticmethod
    def _first_value(node: Any, keys: tuple[str, ...]) -> str | None:
        for k in keys:
            child = node.find(k)
            if child is not None and child.text:
                return child.text.strip()
        return None

    def _first_float(self, node: Any, keys: tuple[str, ...]) -> float | None:
        raw = self._first_value(node, keys)
        if raw is None:
            return None
        # TR formatı: 1.234,56 — virgülü noktaya, binlik nokta sil
        cleaned = raw.replace(".", "").replace(",", ".") if "," in raw else raw
        try:
            return float(cleaned)
        except ValueError:
            return None

    @staticmethod
    def _role_from_type(type_code: str) -> str:
        """1 → customer, 2 → supplier, 3 → both, default both."""
        type_code = (type_code or "").strip().lower()
        if type_code in ("1", "musteri", "customer"):
            return "customer"
        if type_code in ("2", "tedarikci", "supplier"):
            return "supplier"
        return "both"

    @staticmethod
    def _direction_from_type(type_code: str) -> str:
        """1 → outgoing (satış), 2 → incoming (alış)."""
        type_code = (type_code or "").strip().lower()
        if type_code in ("2", "alis", "incoming", "alış"):
            return "incoming"
        return "outgoing"

    @staticmethod
    def _clean_tax_number(raw: str | None) -> str | None:
        if not raw:
            return None
        digits = re.sub(r"\D", "", raw)
        if len(digits) not in (10, 11):  # VKN 10, TCKN 11
            return raw.strip()
        return digits

    @staticmethod
    def _clean_iban(raw: str | None) -> str | None:
        if not raw:
            return None
        s = re.sub(r"\s+", "", raw).upper()
        if s.startswith("TR") and len(s) == 26 and s[2:].isdigit():
            return s
        return s or None

    @staticmethod
    def _normalize_date(raw: str | None) -> str | None:
        if not raw:
            return None
        raw = raw.strip()
        # YYYY-MM-DD
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
            return raw
        # DD.MM.YYYY
        m = re.fullmatch(r"(\d{2})\.(\d{2})\.(\d{4})", raw)
        if m:
            d, mo, y = m.groups()
            return f"{y}-{mo}-{d}"
        # DD/MM/YYYY
        m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4})", raw)
        if m:
            d, mo, y = m.groups()
            return f"{y}-{mo}-{d}"
        return None

    # ── Excel mode (lazy import openpyxl) ──────────────────────────────

    def _parse_excel(self, data: bytes) -> ParseResult:
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            result = ParseResult()
            result.errors.append(ParseError(
                row_index=0,
                record_type="root",
                error_code="missing_dependency",
                error_message="openpyxl yüklü değil. pip install openpyxl gerekli.",
            ))
            return result

        result = ParseResult()
        try:
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            result.errors.append(ParseError(
                row_index=0,
                record_type="root",
                error_code="invalid_xlsx",
                error_message=f"Excel parse hatası: {exc}",
            ))
            return result

        result.source_info = {"sheet_names": list(wb.sheetnames)}

        # Cariler sheet
        cari_sheet = self._find_sheet(wb, ("Cariler", "CARILER", "Customers"))
        if cari_sheet:
            for idx, row in enumerate(_iter_excel_rows(cari_sheet)):
                try:
                    cust = self._customer_from_dict(row)
                    if cust:
                        result.customers.append(cust)
                except (ValueError, TypeError) as exc:
                    result.errors.append(ParseError(
                        row_index=idx,
                        record_type="customer",
                        error_code="invalid_customer",
                        error_message=str(exc),
                    ))

        # Faturalar sheet
        fatura_sheet = self._find_sheet(wb, ("Faturalar", "FATURALAR", "Invoices"))
        if fatura_sheet:
            for idx, row in enumerate(_iter_excel_rows(fatura_sheet)):
                try:
                    inv = self._invoice_from_dict(row)
                    if inv:
                        result.invoices.append(inv)
                except (ValueError, TypeError) as exc:
                    result.errors.append(ParseError(
                        row_index=idx,
                        record_type="invoice",
                        error_code="invalid_invoice",
                        error_message=str(exc),
                    ))

        wb.close()
        return result

    @staticmethod
    def _find_sheet(wb: Any, names: tuple[str, ...]) -> Any | None:
        for n in names:
            if n in wb.sheetnames:
                return wb[n]
        return None

    def _customer_from_dict(self, row: dict[str, Any]) -> ParsedCustomer | None:
        code = _first_dict_value(row, CUSTOMER_KEYS["code"])
        name = _first_dict_value(row, CUSTOMER_KEYS["name"])
        if not code or not name:
            raise ValueError("Cari kodu veya unvan eksik")
        tax_number = _first_dict_value(row, CUSTOMER_KEYS["tax_number"])
        return ParsedCustomer(
            source_code=str(code).strip(),
            name=str(name).strip(),
            tax_number=self._clean_tax_number(tax_number),
            tax_office=_first_dict_value(row, CUSTOMER_KEYS["tax_office"]),
            role=self._role_from_type(
                str(_first_dict_value(row, CUSTOMER_KEYS["type"]) or "")
            ),
            address=_first_dict_value(row, CUSTOMER_KEYS["address"]),
            email=_first_dict_value(row, CUSTOMER_KEYS["email"]),
            phone=_first_dict_value(row, CUSTOMER_KEYS["phone"]),
            iban=self._clean_iban(_first_dict_value(row, CUSTOMER_KEYS["iban"])),
            balance=_safe_float(_first_dict_value(row, CUSTOMER_KEYS["balance"])),
            currency=str(
                _first_dict_value(row, CUSTOMER_KEYS["currency"]) or "TRY"
            ).upper(),
            signature_hash=_signature_hash(
                str(code), str(tax_number or ""), "customer",
            ),
        )

    def _invoice_from_dict(self, row: dict[str, Any]) -> ParsedInvoice | None:
        no = _first_dict_value(row, INVOICE_KEYS["no"])
        cari = _first_dict_value(row, INVOICE_KEYS["customer"])
        issue = _first_dict_value(row, INVOICE_KEYS["issue_date"])
        if not no or not cari or not issue:
            raise ValueError("Fatura no / cari kodu / tarih eksik")
        issue_clean = self._normalize_date(str(issue))
        if not issue_clean:
            raise ValueError(f"Geçersiz tarih: {issue!r}")
        return ParsedInvoice(
            source_no=str(no).strip(),
            customer_source_code=str(cari).strip(),
            issue_date=issue_clean,
            due_date=self._normalize_date(
                str(_first_dict_value(row, INVOICE_KEYS["due_date"]) or "")
            ),
            total_excl_tax=_safe_float(_first_dict_value(row, INVOICE_KEYS["net"])),
            tax_amount=_safe_float(_first_dict_value(row, INVOICE_KEYS["tax"])),
            total_incl_tax=_safe_float(_first_dict_value(row, INVOICE_KEYS["gross"])),
            currency=str(
                _first_dict_value(row, INVOICE_KEYS["currency"]) or "TRY"
            ).upper(),
            direction=self._direction_from_type(
                str(_first_dict_value(row, INVOICE_KEYS["type"]) or "")
            ),
            description=_first_dict_value(row, INVOICE_KEYS["description"]),
            signature_hash=_signature_hash(
                str(no),
                self._direction_from_type(
                    str(_first_dict_value(row, INVOICE_KEYS["type"]) or "")
                ),
                str(cari),
            ),
        )


# ── Helpers ────────────────────────────────────────────────────────────


def _node_to_str(node: Any) -> str:
    """Hata raporu için node'un kısa string temsili."""
    if _ETREE_AVAILABLE and ET is not None and hasattr(ET, "tostring"):
        try:
            return ET.tostring(node, encoding="unicode")  # type: ignore[arg-type]
        except (TypeError, Exception):
            pass
    return f"<{getattr(node, 'tag', '?')}>"


def _signature_hash(*parts: str) -> str:
    raw = "|".join(p for p in parts if p)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:32]


def _first_dict_value(d: dict[str, Any], keys: tuple[str, ...]) -> str | None:
    for k in keys:
        if k in d and d[k] is not None:
            v = d[k]
            return v if isinstance(v, str) else str(v)
    return None


def _safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    cleaned = s.replace(".", "").replace(",", ".") if "," in s else s
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _iter_excel_rows(sheet: Any) -> list[dict[str, Any]]:
    """Excel sheet → list of {header: value} dicts. İlk satır header."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    return out
