"""I3: Mikro ERP connector — XML + Excel modes.

## Mikro ERP

Türkiye'de Logo'dan sonra en yaygın ERP (KOBİ pazarının %25'i).
Mikro'nun "Veri Aktarım Sihirbazı > XML" feature'u Logo'dan
farklı bir tag konvansiyonu kullanır:

  <MIKROIO>
    <CARILER>
      <CARI>
        <KOD>CR001</KOD>
        <UNVAN>Acme Ticaret Ltd.</UNVAN>
        <VERGI_NO>1234567890</VERGI_NO>
        <VERGI_DAIRE>Kadıköy</VERGI_DAIRE>
        <ADRES>...</ADRES>
        <TEL>...</TEL>
        <EMAIL>...</EMAIL>
        <IBAN>...</IBAN>
        <BAKIYE>15000.50</BAKIYE>
        <DOVIZ>TRY</DOVIZ>
        <TIP>ALICI</TIP>          <!-- ALICI / SATICI / KARMA -->
      </CARI>
    </CARILER>
    <FATURALAR>
      <FATURA>
        <NUMARA>F-2026-0001</NUMARA>
        <CARI_KODU>CR001</CARI_KODU>
        <TARIH>2026-05-15</TARIH>
        <VADE>2026-06-15</VADE>
        <ARA_TOPLAM>10000.00</ARA_TOPLAM>
        <KDV_TOPLAM>1800.00</KDV_TOPLAM>
        <GENEL_TOPLAM>11800.00</GENEL_TOPLAM>
        <DOVIZ>TRY</DOVIZ>
        <TUR>SATIS</TUR>           <!-- SATIS / ALIS -->
        <ACIKLAMA>...</ACIKLAMA>
      </FATURA>
    </FATURALAR>
  </MIKROIO>

I1'in framework'ünü kullanır — yeni format, aynı altyapı.
Pure-Python XML parser otomatik fallback (Homebrew Python uyumluluğu).
"""
from __future__ import annotations

import re
from typing import Any

from app.connectors.base import (
    BaseConnector,
    ConnectorMode,
    ParsedCustomer,
    ParsedInvoice,
    ParseError,
    ParseResult,
)
from app.connectors.logo_tiger import (
    _ETREE_AVAILABLE,
    _node_to_str,
    _parse_xml_pure,
    _safe_float,
    _signature_hash,
    ET,
)


# Mikro-specific tag aliases (alternatif Mikro versiyonları için)
CUSTOMER_KEYS: dict[str, tuple[str, ...]] = {
    "code":       ("KOD", "CARI_KODU", "CODE"),
    "name":       ("UNVAN", "AD_SOYAD", "AD", "NAME"),
    "tax_number": ("VERGI_NO", "VKN", "TCKN"),
    "tax_office": ("VERGI_DAIRE", "VERGI_DAIRESI"),
    "address":    ("ADRES", "ADDRESS"),
    "phone":      ("TEL", "TELEFON", "PHONE"),
    "email":      ("EMAIL", "EPOSTA"),
    "iban":       ("IBAN",),
    "balance":    ("BAKIYE", "BALANCE"),
    "currency":   ("DOVIZ", "PARA_BIRIMI", "CURRENCY"),
    "type":       ("TIP", "CARI_TIPI", "TYPE"),
}

INVOICE_KEYS: dict[str, tuple[str, ...]] = {
    "no":          ("NUMARA", "FATURA_NO", "NO"),
    "customer":    ("CARI_KODU", "FATURA_CARI_KODU"),
    "issue_date":  ("TARIH", "FATURA_TARIHI", "DATE"),
    "due_date":    ("VADE", "FATURA_VADE", "DUE_DATE"),
    "net":         ("ARA_TOPLAM", "NET", "TUTAR"),
    "tax":         ("KDV_TOPLAM", "KDV", "VAT"),
    "gross":       ("GENEL_TOPLAM", "BRUT", "TOPLAM"),
    "currency":    ("DOVIZ", "PARA_BIRIMI", "CURRENCY"),
    "type":        ("TUR", "FATURA_TURU", "TYPE"),
    "description": ("ACIKLAMA", "DESCRIPTION"),
}


class MikroConnector(BaseConnector):
    """Mikro ERP connector — XML + Excel modes.

    Logo connector'la aynı framework'ü kullanır; tek fark tag isimleri
    ve TR finansal alan adlandırması (TUR vs FATURA_TIPI gibi).
    """

    connector_type = "mikro"
    supported_modes = (ConnectorMode.XML, ConnectorMode.EXCEL)

    def parse(
        self,
        *,
        data: bytes,
        mode: ConnectorMode,
        filename: str | None = None,
    ) -> ParseResult:
        if mode == ConnectorMode.XML:
            return self._parse_xml(data)
        if mode == ConnectorMode.EXCEL:
            return self._parse_excel(data)
        if mode == ConnectorMode.WEB_SERVICE:
            raise NotImplementedError(
                "Mikro Web Service modu henüz aktif değil — resmi API anlaşması gerekli."
            )
        raise ValueError(f"Desteklenmeyen mod: {mode}")

    # ── XML mode ───────────────────────────────────────────────────────

    def _parse_xml(self, data: bytes) -> ParseResult:
        result = ParseResult()
        root: Any = None
        if _ETREE_AVAILABLE:
            try:
                root = ET.fromstring(data)
            except Exception:
                root = None
        if root is None:
            try:
                root = _parse_xml_pure(data)
            except (ValueError, Exception) as exc:
                result.errors.append(ParseError(
                    row_index=0,
                    record_type="root",
                    error_code="invalid_xml",
                    error_message=f"XML parse hatası: {exc}",
                ))
                return result

        result.source_info = {
            "root_tag": root.tag,
            "child_count": len(list(root)),
            "vendor": "mikro",
        }

        # Mikro: <CARILER>/<CARI> + <FATURALAR>/<FATURA>
        cari_nodes = self._collect_nodes(root, ("CARI", "CUSTOMER"))
        for idx, node in enumerate(cari_nodes):
            try:
                customer = self._parse_customer_node(node)
                if customer:
                    result.customers.append(customer)
            except (ValueError, TypeError) as exc:
                result.errors.append(ParseError(
                    row_index=idx,
                    record_type="customer",
                    error_code="invalid_customer",
                    error_message=str(exc),
                    raw_payload=_node_to_str(node)[:500],
                ))

        fatura_nodes = self._collect_nodes(root, ("FATURA", "INVOICE"))
        for idx, node in enumerate(fatura_nodes):
            try:
                invoice = self._parse_invoice_node(node)
                if invoice:
                    result.invoices.append(invoice)
            except (ValueError, TypeError) as exc:
                result.errors.append(ParseError(
                    row_index=idx,
                    record_type="invoice",
                    error_code="invalid_invoice",
                    error_message=str(exc),
                    raw_payload=_node_to_str(node)[:500],
                ))

        return result

    @staticmethod
    def _collect_nodes(
        root: Any, tags: tuple[str, ...]
    ) -> list[Any]:
        nodes: list[Any] = []
        for tag in tags:
            nodes.extend(root.findall(f".//{tag}"))
        seen_ids: set[int] = set()
        unique: list[Any] = []
        for n in nodes:
            nid = id(n)
            if nid not in seen_ids:
                seen_ids.add(nid)
                unique.append(n)
        return unique

    def _parse_customer_node(self, node: Any) -> ParsedCustomer | None:
        code = self._first_value(node, CUSTOMER_KEYS["code"])
        name = self._first_value(node, CUSTOMER_KEYS["name"])
        if not code or not name:
            raise ValueError("Cari kodu veya unvan eksik")

        tax_number = self._first_value(node, CUSTOMER_KEYS["tax_number"])
        type_code = self._first_value(node, CUSTOMER_KEYS["type"]) or ""

        return ParsedCustomer(
            source_code=code.strip(),
            name=name.strip(),
            tax_number=self._clean_tax_number(tax_number),
            tax_office=(
                self._first_value(node, CUSTOMER_KEYS["tax_office"]) or None
            ),
            role=self._role_from_type(type_code),
            address=self._first_value(node, CUSTOMER_KEYS["address"]),
            email=self._first_value(node, CUSTOMER_KEYS["email"]),
            phone=self._first_value(node, CUSTOMER_KEYS["phone"]),
            iban=self._clean_iban(self._first_value(node, CUSTOMER_KEYS["iban"])),
            balance=self._first_float(node, CUSTOMER_KEYS["balance"]) or 0.0,
            currency=(
                self._first_value(node, CUSTOMER_KEYS["currency"]) or "TRY"
            ).upper(),
            signature_hash=_signature_hash(
                code, tax_number or "", "mikro_customer",
            ),
        )

    def _parse_invoice_node(self, node: Any) -> ParsedInvoice | None:
        no = self._first_value(node, INVOICE_KEYS["no"])
        cari = self._first_value(node, INVOICE_KEYS["customer"])
        issue = self._first_value(node, INVOICE_KEYS["issue_date"])
        if not no or not cari or not issue:
            raise ValueError("Fatura no / cari kodu / tarih eksik")

        issue_clean = self._normalize_date(issue)
        if not issue_clean:
            raise ValueError(f"Geçersiz tarih: {issue!r}")

        due = self._first_value(node, INVOICE_KEYS["due_date"])
        due_clean = self._normalize_date(due) if due else None

        net = self._first_float(node, INVOICE_KEYS["net"]) or 0.0
        tax = self._first_float(node, INVOICE_KEYS["tax"]) or 0.0
        gross = self._first_float(node, INVOICE_KEYS["gross"]) or (net + tax)

        type_code = self._first_value(node, INVOICE_KEYS["type"]) or ""
        direction = self._direction_from_type(type_code)

        return ParsedInvoice(
            source_no=no.strip(),
            customer_source_code=cari.strip(),
            issue_date=issue_clean,
            due_date=due_clean,
            total_excl_tax=net,
            tax_amount=tax,
            total_incl_tax=gross,
            currency=(
                self._first_value(node, INVOICE_KEYS["currency"]) or "TRY"
            ).upper(),
            direction=direction,
            description=self._first_value(node, INVOICE_KEYS["description"]),
            signature_hash=_signature_hash(no, direction, cari),
        )

    # ── Helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _first_value(node: Any, keys: tuple[str, ...]) -> str | None:
        for k in keys:
            child = node.find(k)
            if child is not None and child.text:
                text: str = child.text.strip()
                return text
        return None

    def _first_float(self, node: Any, keys: tuple[str, ...]) -> float | None:
        raw = self._first_value(node, keys)
        if raw is None:
            return None
        return _safe_float(raw)

    @staticmethod
    def _role_from_type(type_code: str) -> str:
        """Mikro: ALICI=customer, SATICI=supplier, KARMA=both."""
        t = (type_code or "").strip().upper()
        if t in ("ALICI", "MUSTERI", "CUSTOMER", "1"):
            return "customer"
        if t in ("SATICI", "TEDARIKCI", "SUPPLIER", "2"):
            return "supplier"
        return "both"

    @staticmethod
    def _direction_from_type(type_code: str) -> str:
        """Mikro: SATIS=outgoing, ALIS=incoming."""
        t = (type_code or "").strip().upper()
        if t in ("ALIS", "INCOMING", "ALIŞ", "2"):
            return "incoming"
        return "outgoing"

    @staticmethod
    def _clean_tax_number(raw: str | None) -> str | None:
        if not raw:
            return None
        digits = re.sub(r"\D", "", raw)
        if len(digits) not in (10, 11):
            return raw.strip()
        return digits

    @staticmethod
    def _clean_iban(raw: str | None) -> str | None:
        if not raw:
            return None
        s = re.sub(r"\s+", "", raw).upper()
        if s.startswith("TR") and len(s) == 26 and s[2:].isdigit():
            return s
        return s or None

    @staticmethod
    def _normalize_date(raw: str | None) -> str | None:
        if not raw:
            return None
        raw = raw.strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
            return raw
        m = re.fullmatch(r"(\d{2})\.(\d{2})\.(\d{4})", raw)
        if m:
            d, mo, y = m.groups()
            return f"{y}-{mo}-{d}"
        m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4})", raw)
        if m:
            d, mo, y = m.groups()
            return f"{y}-{mo}-{d}"
        return None

    # ── Excel mode ─────────────────────────────────────────────────────

    def _parse_excel(self, data: bytes) -> ParseResult:
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            result = ParseResult()
            result.errors.append(ParseError(
                row_index=0, record_type="root",
                error_code="missing_dependency",
                error_message="openpyxl yüklü değil. pip install openpyxl gerekli.",
            ))
            return result

        result = ParseResult()
        try:
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            result.errors.append(ParseError(
                row_index=0, record_type="root",
                error_code="invalid_xlsx",
                error_message=f"Excel parse hatası: {exc}",
            ))
            return result

        result.source_info = {
            "sheet_names": list(wb.sheetnames),
            "vendor": "mikro",
        }

        # Mikro Excel: "Cariler" + "Faturalar" veya "CARI" + "FATURA"
        cari_sheet = self._find_sheet(
            wb, ("Cariler", "CARILER", "CARI", "Customers"),
        )
        if cari_sheet:
            for idx, row in enumerate(_iter_excel_rows(cari_sheet)):
                try:
                    cust = self._customer_from_dict(row)
                    if cust:
                        result.customers.append(cust)
                except (ValueError, TypeError) as exc:
                    result.errors.append(ParseError(
                        row_index=idx, record_type="customer",
                        error_code="invalid_customer", error_message=str(exc),
                    ))

        fatura_sheet = self._find_sheet(
            wb, ("Faturalar", "FATURALAR", "FATURA", "Invoices"),
        )
        if fatura_sheet:
            for idx, row in enumerate(_iter_excel_rows(fatura_sheet)):
                try:
                    inv = self._invoice_from_dict(row)
                    if inv:
                        result.invoices.append(inv)
                except (ValueError, TypeError) as exc:
                    result.errors.append(ParseError(
                        row_index=idx, record_type="invoice",
                        error_code="invalid_invoice", error_message=str(exc),
                    ))

        wb.close()
        return result

    @staticmethod
    def _find_sheet(wb: Any, names: tuple[str, ...]) -> Any | None:
        for n in names:
            if n in wb.sheetnames:
                return wb[n]
        return None

    def _customer_from_dict(self, row: dict[str, Any]) -> ParsedCustomer | None:
        code = _first_dict_value(row, CUSTOMER_KEYS["code"])
        name = _first_dict_value(row, CUSTOMER_KEYS["name"])
        if not code or not name:
            raise ValueError("Cari kodu veya unvan eksik")
        tax_number = _first_dict_value(row, CUSTOMER_KEYS["tax_number"])
        return ParsedCustomer(
            source_code=str(code).strip(),
            name=str(name).strip(),
            tax_number=self._clean_tax_number(tax_number),
            tax_office=_first_dict_value(row, CUSTOMER_KEYS["tax_office"]),
            role=self._role_from_type(
                str(_first_dict_value(row, CUSTOMER_KEYS["type"]) or "")
            ),
            address=_first_dict_value(row, CUSTOMER_KEYS["address"]),
            email=_first_dict_value(row, CUSTOMER_KEYS["email"]),
            phone=_first_dict_value(row, CUSTOMER_KEYS["phone"]),
            iban=self._clean_iban(_first_dict_value(row, CUSTOMER_KEYS["iban"])),
            balance=_safe_float(_first_dict_value(row, CUSTOMER_KEYS["balance"])),
            currency=str(
                _first_dict_value(row, CUSTOMER_KEYS["currency"]) or "TRY"
            ).upper(),
            signature_hash=_signature_hash(
                str(code), str(tax_number or ""), "mikro_customer",
            ),
        )

    def _invoice_from_dict(self, row: dict[str, Any]) -> ParsedInvoice | None:
        no = _first_dict_value(row, INVOICE_KEYS["no"])
        cari = _first_dict_value(row, INVOICE_KEYS["customer"])
        issue = _first_dict_value(row, INVOICE_KEYS["issue_date"])
        if not no or not cari or not issue:
            raise ValueError("Fatura no / cari kodu / tarih eksik")
        issue_clean = self._normalize_date(str(issue))
        if not issue_clean:
            raise ValueError(f"Geçersiz tarih: {issue!r}")
        type_code = str(_first_dict_value(row, INVOICE_KEYS["type"]) or "")
        direction = self._direction_from_type(type_code)
        return ParsedInvoice(
            source_no=str(no).strip(),
            customer_source_code=str(cari).strip(),
            issue_date=issue_clean,
            due_date=self._normalize_date(
                str(_first_dict_value(row, INVOICE_KEYS["due_date"]) or "")
            ),
            total_excl_tax=_safe_float(_first_dict_value(row, INVOICE_KEYS["net"])),
            tax_amount=_safe_float(_first_dict_value(row, INVOICE_KEYS["tax"])),
            total_incl_tax=_safe_float(
                _first_dict_value(row, INVOICE_KEYS["gross"])
            ),
            currency=str(
                _first_dict_value(row, INVOICE_KEYS["currency"]) or "TRY"
            ).upper(),
            direction=direction,
            description=_first_dict_value(row, INVOICE_KEYS["description"]),
            signature_hash=_signature_hash(str(no), direction, str(cari)),
        )


# ── Module-level helpers (Logo'dan import edilemeyen küçük utils) ──────


def _first_dict_value(d: dict[str, Any], keys: tuple[str, ...]) -> str | None:
    for k in keys:
        if k in d and d[k] is not None:
            v = d[k]
            return v if isinstance(v, str) else str(v)
    return None


def _iter_excel_rows(sheet: Any) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    return out
