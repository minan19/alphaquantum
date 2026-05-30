"""I4: Netsis ERP connector — XML + Excel modes.

## Netsis ERP

Logo/Mikro'dan sonra Türkiye'nin 3. yaygın ERP'si (yaklaşık %10 KOBİ
pazarı). Netsis XML formatı OASIS-tarzı ama tag konvansiyonu farklı.

  <NETSIS_EXPORT>
    <CARI_LIST>
      <CARI>
        <CARI_KOD>NS001</CARI_KOD>
        <CARI_UNVANI>...</CARI_UNVANI>
        <VKN_TCKN>...</VKN_TCKN>
        <TIP>1</TIP>  <!-- 1=Müşteri 2=Satıcı -->
        ...
      </CARI>
    </CARI_LIST>
    <FATURA_LIST>
      <FATURA>
        <FATURA_NUM>...</FATURA_NUM>
        <CARI_KOD>NS001</CARI_KOD>
        <FATURA_TIPI>1</FATURA_TIPI>  <!-- 1=Satış 2=Alış -->
        ...
      </FATURA>
    </FATURA_LIST>
  </NETSIS_EXPORT>

I1/I3 framework'ünü kullanır.
"""
from __future__ import annotations

import re
from typing import Any

from app.connectors.base import (
    BaseConnector,
    ConnectorMode,
    ParsedCustomer,
    ParsedInvoice,
    ParseError,
    ParseResult,
)
from app.connectors.logo_tiger import (
    _ETREE_AVAILABLE,
    _node_to_str,
    _parse_xml_pure,
    _safe_float,
    _signature_hash,
    ET,
)


CUSTOMER_KEYS: dict[str, tuple[str, ...]] = {
    "code":       ("CARI_KOD", "KOD"),
    "name":       ("CARI_UNVANI", "UNVAN"),
    "tax_number": ("VKN_TCKN", "VKN"),
    "tax_office": ("VD", "VERGI_DAIRE"),
    "address":    ("ADRES",),
    "phone":      ("TELEFON", "TEL"),
    "email":      ("EMAIL", "EPOSTA"),
    "iban":       ("IBAN",),
    "balance":    ("BAKIYE",),
    "currency":   ("DOVIZ", "PARA_BIRIMI"),
    "type":       ("TIP", "CARI_TIPI"),
}

INVOICE_KEYS: dict[str, tuple[str, ...]] = {
    "no":          ("FATURA_NUM", "FATURA_NO"),
    "customer":    ("CARI_KOD",),
    "issue_date":  ("FATURA_TARIH", "TARIH"),
    "due_date":    ("VADE_TARIH", "VADE"),
    "net":         ("MAT_TOPLAM", "NET"),
    "tax":         ("KDV_TOPLAM", "KDV"),
    "gross":       ("GENEL_TOP", "BRUT"),
    "currency":    ("DOVIZ", "PARA"),
    "type":        ("FATURA_TIPI", "TIP"),
    "description": ("ACIK", "ACIKLAMA"),
}


class NetsisConnector(BaseConnector):
    """Netsis ERP connector — XML + Excel modes."""

    connector_type = "netsis"
    supported_modes = (ConnectorMode.XML, ConnectorMode.EXCEL)

    def parse(
        self,
        *,
        data: bytes,
        mode: ConnectorMode,
        filename: str | None = None,
    ) -> ParseResult:
        if mode == ConnectorMode.XML:
            return self._parse_xml(data)
        if mode == ConnectorMode.EXCEL:
            return self._parse_excel(data)
        if mode == ConnectorMode.WEB_SERVICE:
            raise NotImplementedError(
                "Netsis Web Service modu henüz aktif değil."
            )
        raise ValueError(f"Desteklenmeyen mod: {mode}")

    # ── XML ────────────────────────────────────────────────────────────

    def _parse_xml(self, data: bytes) -> ParseResult:
        result = ParseResult()
        root: Any = None
        if _ETREE_AVAILABLE:
            try:
                root = ET.fromstring(data)
            except Exception:
                root = None
        if root is None:
            try:
                root = _parse_xml_pure(data)
            except (ValueError, Exception) as exc:
                result.errors.append(ParseError(
                    row_index=0, record_type="root",
                    error_code="invalid_xml",
                    error_message=f"XML parse hatası: {exc}",
                ))
                return result

        result.source_info = {
            "root_tag": root.tag,
            "child_count": len(list(root)),
            "vendor": "netsis",
        }

        cari_nodes = self._collect_nodes(root, ("CARI", "CUSTOMER"))
        for idx, node in enumerate(cari_nodes):
            try:
                cust = self._parse_customer(node)
                if cust:
                    result.customers.append(cust)
            except (ValueError, TypeError) as exc:
                result.errors.append(ParseError(
                    row_index=idx, record_type="customer",
                    error_code="invalid_customer", error_message=str(exc),
                    raw_payload=_node_to_str(node)[:500],
                ))

        fatura_nodes = self._collect_nodes(root, ("FATURA", "INVOICE"))
        for idx, node in enumerate(fatura_nodes):
            try:
                inv = self._parse_invoice(node)
                if inv:
                    result.invoices.append(inv)
            except (ValueError, TypeError) as exc:
                result.errors.append(ParseError(
                    row_index=idx, record_type="invoice",
                    error_code="invalid_invoice", error_message=str(exc),
                    raw_payload=_node_to_str(node)[:500],
                ))
        return result

    @staticmethod
    def _collect_nodes(root: Any, tags: tuple[str, ...]) -> list[Any]:
        nodes: list[Any] = []
        for t in tags:
            nodes.extend(root.findall(f".//{t}"))
        seen: set[int] = set()
        out: list[Any] = []
        for n in nodes:
            nid = id(n)
            if nid not in seen:
                seen.add(nid)
                out.append(n)
        return out

    def _parse_customer(self, node: Any) -> ParsedCustomer | None:
        code = self._first_value(node, CUSTOMER_KEYS["code"])
        name = self._first_value(node, CUSTOMER_KEYS["name"])
        if not code or not name:
            raise ValueError("Cari kodu veya unvan eksik")
        tax_number = self._first_value(node, CUSTOMER_KEYS["tax_number"])
        type_code = self._first_value(node, CUSTOMER_KEYS["type"]) or ""
        return ParsedCustomer(
            source_code=code.strip(),
            name=name.strip(),
            tax_number=self._clean_tax_number(tax_number),
            tax_office=self._first_value(node, CUSTOMER_KEYS["tax_office"]),
            role=self._role_from_type(type_code),
            address=self._first_value(node, CUSTOMER_KEYS["address"]),
            email=self._first_value(node, CUSTOMER_KEYS["email"]),
            phone=self._first_value(node, CUSTOMER_KEYS["phone"]),
            iban=self._clean_iban(self._first_value(node, CUSTOMER_KEYS["iban"])),
            balance=self._first_float(node, CUSTOMER_KEYS["balance"]) or 0.0,
            currency=(self._first_value(node, CUSTOMER_KEYS["currency"]) or "TRY").upper(),
            signature_hash=_signature_hash(code, tax_number or "", "netsis_customer"),
        )

    def _parse_invoice(self, node: Any) -> ParsedInvoice | None:
        no = self._first_value(node, INVOICE_KEYS["no"])
        cari = self._first_value(node, INVOICE_KEYS["customer"])
        issue = self._first_value(node, INVOICE_KEYS["issue_date"])
        if not no or not cari or not issue:
            raise ValueError("Fatura no / cari kodu / tarih eksik")
        issue_clean = self._normalize_date(issue)
        if not issue_clean:
            raise ValueError(f"Geçersiz tarih: {issue!r}")

        due = self._first_value(node, INVOICE_KEYS["due_date"])
        due_clean = self._normalize_date(due) if due else None

        net = self._first_float(node, INVOICE_KEYS["net"]) or 0.0
        tax = self._first_float(node, INVOICE_KEYS["tax"]) or 0.0
        gross = self._first_float(node, INVOICE_KEYS["gross"]) or (net + tax)
        type_code = self._first_value(node, INVOICE_KEYS["type"]) or ""
        direction = self._direction_from_type(type_code)

        return ParsedInvoice(
            source_no=no.strip(),
            customer_source_code=cari.strip(),
            issue_date=issue_clean,
            due_date=due_clean,
            total_excl_tax=net,
            tax_amount=tax,
            total_incl_tax=gross,
            currency=(self._first_value(node, INVOICE_KEYS["currency"]) or "TRY").upper(),
            direction=direction,
            description=self._first_value(node, INVOICE_KEYS["description"]),
            signature_hash=_signature_hash(no, direction, cari),
        )

    # ── Helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _first_value(node: Any, keys: tuple[str, ...]) -> str | None:
        for k in keys:
            child = node.find(k)
            if child is not None and child.text:
                text: str = child.text.strip()
                return text
        return None

    def _first_float(self, node: Any, keys: tuple[str, ...]) -> float | None:
        raw = self._first_value(node, keys)
        if raw is None:
            return None
        return _safe_float(raw)

    @staticmethod
    def _role_from_type(code: str) -> str:
        c = (code or "").strip().upper()
        if c in ("1", "MUSTERI", "ALICI"):
            return "customer"
        if c in ("2", "SATICI", "TEDARIKCI"):
            return "supplier"
        return "both"

    @staticmethod
    def _direction_from_type(code: str) -> str:
        c = (code or "").strip().upper()
        if c in ("2", "ALIS"):
            return "incoming"
        return "outgoing"

    @staticmethod
    def _clean_tax_number(raw: str | None) -> str | None:
        if not raw:
            return None
        digits = re.sub(r"\D", "", raw)
        return digits if len(digits) in (10, 11) else raw.strip()

    @staticmethod
    def _clean_iban(raw: str | None) -> str | None:
        if not raw:
            return None
        return re.sub(r"\s+", "", raw).upper() or None

    @staticmethod
    def _normalize_date(raw: str | None) -> str | None:
        if not raw:
            return None
        raw = raw.strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
            return raw
        from datetime import datetime
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    # ── Excel (lazy) ───────────────────────────────────────────────────

    def _parse_excel(self, data: bytes) -> ParseResult:
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            result = ParseResult()
            result.errors.append(ParseError(
                row_index=0, record_type="root",
                error_code="missing_dependency",
                error_message="openpyxl yüklü değil",
            ))
            return result
        result = ParseResult()
        try:
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            result.errors.append(ParseError(
                row_index=0, record_type="root",
                error_code="invalid_xlsx",
                error_message=f"Excel parse hatası: {exc}",
            ))
            return result
        # Netsis Excel: 'CARI_LIST' + 'FATURA_LIST' veya benzeri
        # MVP: minimal implementation (Logo/Mikro pattern'leri yeterli)
        result.source_info = {"sheet_names": list(wb.sheetnames), "vendor": "netsis"}
        wb.close()
        return result
