from __future__ import annotations

import hashlib
import hmac
import io
from typing import Any


class ReportingEngine:
    """Generate signed PDF and Excel exports for finance data."""

    # ── Signing ────────────────────────────────────────────────────────────────

    @staticmethod
    def sign(content: bytes, secret: str) -> str:
        """Return HMAC-SHA256 hex digest of *content* keyed with *secret*."""
        digest = hmac.new(
            secret.encode("utf-8"),
            content,
            hashlib.sha256,
        ).hexdigest()
        return f"hmac-sha256={digest}"

    # ── Excel exports ──────────────────────────────────────────────────────────

    @staticmethod
    def ledger_to_xlsx(entries: list[dict[str, Any]]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Ledger"

        headers = ["ID", "Company", "Type", "Amount", "Category", "Description", "Date", "Created At"]
        ws.append(headers)

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for entry in entries:
            ws.append([
                entry.get("id"),
                entry.get("company_name") or entry.get("company"),
                entry.get("entry_type"),
                entry.get("amount"),
                entry.get("category"),
                entry.get("description"),
                entry.get("entry_date"),
                entry.get("created_at"),
            ])

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 2, 50))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def budget_vs_actual_to_xlsx(
        *,
        company: str | None,
        year: int,
        month: int | None,
        items: list[dict[str, Any]],
        totals: dict[str, float],
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Budget vs Actual"

        # Meta section
        ws.append(["Alpha Quantum – Budget vs Actual Report"])
        ws["A1"].font = Font(bold=True, size=14)
        period = f"{year}" if month is None else f"{year}-{month:02d}"
        ws.append(["Period:", period])
        ws.append(["Company:", company or "All"])
        ws.append([])

        # Table headers
        headers = ["Category", "Type", "Budget", "Actual", "Variance", "Variance %", "Status"]
        ws.append(headers)
        header_row = ws.max_row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for item in items:
            ws.append([
                item.get("category"),
                item.get("entry_type"),
                item.get("budget_amount"),
                item.get("actual_amount"),
                item.get("variance"),
                item.get("variance_pct"),
                item.get("status"),
            ])

        # Totals row
        ws.append([])
        ws.append([
            "TOTALS", "",
            totals.get("total_budget_income", 0) - totals.get("total_budget_expense", 0),
            totals.get("total_actual_income", 0) - totals.get("total_actual_expense", 0),
            totals.get("net_variance", 0), "", "",
        ])
        totals_row = ws.max_row
        for cell in ws[totals_row]:
            cell.font = Font(bold=True)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 2, 40))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── PDF exports ────────────────────────────────────────────────────────────

    @staticmethod
    def ledger_to_pdf(entries: list[dict[str, Any]]) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Alpha Quantum – Finance Ledger", styles["Title"]))
        elements.append(Spacer(1, 12))

        headers = [["ID", "Company", "Type", "Amount", "Category", "Description", "Date"]]
        rows = headers + [
            [
                str(e.get("id", "")),
                str(e.get("company_name") or e.get("company") or ""),
                str(e.get("entry_type", "")),
                f"{e.get('amount', 0):.2f}",
                str(e.get("category", "")),
                str(e.get("description") or "")[:40],
                str(e.get("entry_date", "")),
            ]
            for e in entries
        ]

        col_widths = [35, 100, 60, 65, 80, 200, 75]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ]))
        elements.append(table)
        doc.build(elements)
        return buf.getvalue()

    @staticmethod
    def budget_vs_actual_to_pdf(
        *,
        company: str | None,
        year: int,
        month: int | None,
        items: list[dict[str, Any]],
        totals: dict[str, float],
    ) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        period = f"{year}" if month is None else f"{year}-{month:02d}"
        scope = company or "All Companies"
        elements.append(Paragraph("Alpha Quantum – Budget vs Actual", styles["Title"]))
        elements.append(Paragraph(f"Period: {period}  |  Scope: {scope}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        status_colors = {
            "ON_TRACK": colors.HexColor("#1A7A4A"),
            "OVER": colors.HexColor("#C0392B"),
            "UNDER": colors.HexColor("#E67E22"),
        }

        headers = [["Category", "Type", "Budget", "Actual", "Variance", "Var %", "Status"]]
        rows = headers + [
            [
                str(i.get("category", "")),
                str(i.get("entry_type", "")),
                f"{i.get('budget_amount', 0):.2f}",
                f"{i.get('actual_amount', 0):.2f}",
                f"{i.get('variance', 0):+.2f}",
                f"{i.get('variance_pct', 0):+.1f}%",
                str(i.get("status", "")),
            ]
            for i in items
        ]

        table = Table(rows, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
            ("ALIGN", (2, 1), (-2, -1), "RIGHT"),
        ]
        # Colour the Status column per value
        for row_idx, item in enumerate(items, start=1):
            st = str(item.get("status", ""))
            clr = status_colors.get(st, colors.black)
            style_cmds.append(("TEXTCOLOR", (6, row_idx), (6, row_idx), clr))
            style_cmds.append(("FONTNAME", (6, row_idx), (6, row_idx), "Helvetica-Bold"))

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

        # Summary footer
        elements.append(Spacer(1, 16))
        net_budget = totals.get("net_budget", 0)
        net_actual = totals.get("net_actual", 0)
        net_variance = totals.get("net_variance", 0)
        summary_data = [
            ["", "Budget Net", "Actual Net", "Net Variance"],
            ["", f"{net_budget:.2f}", f"{net_actual:.2f}", f"{net_variance:+.2f}"],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.grey),
        ]))
        elements.append(summary_table)
        doc.build(elements)
        return buf.getvalue()

    @staticmethod
    def invoice_to_pdf(
        invoice: dict[str, Any],
        *,
        customer: dict[str, Any] | None = None,
    ) -> bytes:
        """QW-2 — Single-invoice PDF export (customer-facing or archival).

        Renders a one-page document with: invoice header, company/customer
        information, line-amount summary table, and status colour-coding.
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        inv_number = invoice.get("invoice_number") or f"#{invoice.get('id', '')}"
        title = invoice.get("title") or ""
        company = invoice.get("company_name") or invoice.get("company") or ""
        amount = float(invoice.get("amount", 0))
        paid = float(invoice.get("paid_amount") or 0)
        outstanding = round(amount - paid, 2)
        currency = invoice.get("currency") or "TRY"
        status = str(invoice.get("status", "pending")).upper()

        status_palette = {
            "PAID":      colors.HexColor("#1A7A4A"),
            "PARTIAL":   colors.HexColor("#E67E22"),
            "OVERDUE":   colors.HexColor("#C0392B"),
            "PENDING":   colors.HexColor("#1F4E79"),
            "CANCELLED": colors.HexColor("#6C757D"),
        }
        status_color = status_palette.get(status, colors.black)

        elements.append(Paragraph(
            f"Alpha Quantum – Fatura {inv_number}", styles["Title"],
        ))
        elements.append(Paragraph(f"<b>Şirket:</b> {company}", styles["Normal"]))
        if customer is not None:
            cust_name = customer.get("full_name", "")
            cust_email = customer.get("email", "")
            cust_phone = customer.get("phone", "")
            elements.append(Paragraph(
                f"<b>Müşteri:</b> {cust_name}" +
                (f" — {cust_email}" if cust_email else "") +
                (f" / {cust_phone}" if cust_phone else ""),
                styles["Normal"],
            ))
        elements.append(Paragraph(f"<b>Başlık:</b> {title}", styles["Normal"]))
        elements.append(Spacer(1, 8))

        info_data = [
            ["Düzenlenme Tarihi", str(invoice.get("issue_date", ""))],
            ["Vade Tarihi", str(invoice.get("due_date", ""))],
            ["Ödeme Tarihi", str(invoice.get("paid_date") or "—")],
            ["Para Birimi", str(currency)],
            ["Durum", status],
        ]
        info_table = Table(info_data, colWidths=[150, 250])
        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1),
                [colors.white, colors.HexColor("#F4F8FC")]),
            # colour the Status row
            ("TEXTCOLOR", (1, 4), (1, 4), status_color),
            ("FONTNAME", (1, 4), (1, 4), "Helvetica-Bold"),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 16))

        # Money summary
        money_data = [
            ["Toplam Tutar", f"{amount:,.2f} {currency}"],
            ["Ödenen", f"{paid:,.2f} {currency}"],
            ["Açık Bakiye", f"{outstanding:,.2f} {currency}"],
        ]
        money_table = Table(money_data, colWidths=[150, 250])
        money_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LINEABOVE", (0, 2), (-1, 2), 1.0, colors.grey),
            ("LINEBELOW", (0, 2), (-1, 2), 1.5, colors.HexColor("#1F4E79")),
            ("FONTSIZE", (0, 2), (-1, 2), 11),
            ("TEXTCOLOR", (1, 2), (1, 2), status_color),
        ]))
        elements.append(money_table)

        description = (invoice.get("description") or "").strip()
        if description:
            elements.append(Spacer(1, 18))
            elements.append(Paragraph(
                f"<b>Açıklama:</b> {description}", styles["Normal"],
            ))

        doc.build(elements)
        return buf.getvalue()
