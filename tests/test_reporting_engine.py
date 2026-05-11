from __future__ import annotations

import hashlib
import hmac
import io
import unittest

from app.engines.reporting_engine import ReportingEngine

_OPENPYXL_OK = False
try:
    import openpyxl as _openpyxl  # noqa: F401
    _OPENPYXL_OK = True
except Exception:
    pass


SAMPLE_ENTRIES = [
    {
        "id": 1,
        "company_name": "Alpha Corp",
        "entry_type": "income",
        "amount": 5000.0,
        "category": "sales",
        "description": "Q1 revenue",
        "entry_date": "2026-01-15",
        "created_at": 1700000000,
    },
    {
        "id": 2,
        "company_name": "Alpha Corp",
        "entry_type": "expense",
        "amount": 1200.0,
        "category": "utilities",
        "description": "Electricity",
        "entry_date": "2026-01-20",
        "created_at": 1700000100,
    },
]

SAMPLE_ITEMS = [
    {
        "category": "sales",
        "entry_type": "income",
        "budget_amount": 10000.0,
        "actual_amount": 9500.0,
        "variance": -500.0,
        "variance_pct": -5.0,
        "status": "UNDER",
    },
    {
        "category": "utilities",
        "entry_type": "expense",
        "budget_amount": 1500.0,
        "actual_amount": 1200.0,
        "variance": -300.0,
        "variance_pct": -20.0,
        "status": "ON_TRACK",
    },
]

SAMPLE_TOTALS = {
    "total_budget_income": 10000.0,
    "total_budget_expense": 1500.0,
    "total_actual_income": 9500.0,
    "total_actual_expense": 1200.0,
    "net_budget": 8500.0,
    "net_actual": 8300.0,
    "net_variance": -200.0,
}


class ReportingEngineSignTests(unittest.TestCase):
    def test_sign_returns_hmac_sha256_prefix(self):
        engine = ReportingEngine()
        sig = engine.sign(b"hello world", "secret")
        self.assertTrue(sig.startswith("hmac-sha256="))

    def test_sign_correct_digest(self):
        engine = ReportingEngine()
        content = b"test content"
        secret = "my-secret"
        sig = engine.sign(content, secret)
        expected = hmac.new(secret.encode(), content, hashlib.sha256).hexdigest()
        self.assertEqual(sig, f"hmac-sha256={expected}")

    def test_sign_different_secrets_differ(self):
        engine = ReportingEngine()
        content = b"data"
        self.assertNotEqual(engine.sign(content, "key1"), engine.sign(content, "key2"))

    def test_sign_different_content_differs(self):
        engine = ReportingEngine()
        self.assertNotEqual(engine.sign(b"a", "key"), engine.sign(b"b", "key"))


@unittest.skipUnless(_OPENPYXL_OK, "openpyxl unavailable (libexpat ABI mismatch on this machine)")
class ReportingEngineXlsxTests(unittest.TestCase):
    def test_ledger_to_xlsx_returns_bytes(self):
        engine = ReportingEngine()
        result = engine.ledger_to_xlsx(SAMPLE_ENTRIES)
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_ledger_to_xlsx_valid_workbook(self):
        from openpyxl import load_workbook

        engine = ReportingEngine()
        result = engine.ledger_to_xlsx(SAMPLE_ENTRIES)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Header row + 2 data rows
        self.assertEqual(ws.max_row, 3)
        self.assertEqual(ws.cell(1, 1).value, "ID")

    def test_ledger_to_xlsx_empty_entries(self):
        engine = ReportingEngine()
        result = engine.ledger_to_xlsx([])
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_budget_vs_actual_to_xlsx_returns_bytes(self):
        engine = ReportingEngine()
        result = engine.budget_vs_actual_to_xlsx(
            company="Alpha Corp",
            year=2026,
            month=1,
            items=SAMPLE_ITEMS,
            totals=SAMPLE_TOTALS,
        )
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_budget_vs_actual_to_xlsx_valid_workbook(self):
        from openpyxl import load_workbook

        engine = ReportingEngine()
        result = engine.budget_vs_actual_to_xlsx(
            company=None,
            year=2026,
            month=None,
            items=SAMPLE_ITEMS,
            totals=SAMPLE_TOTALS,
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertIsNotNone(ws)
        # Must have data rows beyond just headers
        self.assertGreater(ws.max_row, 5)


class ReportingEnginePdfTests(unittest.TestCase):
    def test_ledger_to_pdf_returns_pdf_bytes(self):
        engine = ReportingEngine()
        result = engine.ledger_to_pdf(SAMPLE_ENTRIES)
        self.assertIsInstance(result, bytes)
        self.assertTrue(result.startswith(b"%PDF"), "Expected PDF magic bytes")

    def test_ledger_to_pdf_empty_entries(self):
        engine = ReportingEngine()
        result = engine.ledger_to_pdf([])
        self.assertIsInstance(result, bytes)
        self.assertTrue(result.startswith(b"%PDF"))

    def test_budget_vs_actual_to_pdf_returns_pdf_bytes(self):
        engine = ReportingEngine()
        result = engine.budget_vs_actual_to_pdf(
            company="Alpha Corp",
            year=2026,
            month=3,
            items=SAMPLE_ITEMS,
            totals=SAMPLE_TOTALS,
        )
        self.assertIsInstance(result, bytes)
        self.assertTrue(result.startswith(b"%PDF"))

    def test_budget_vs_actual_to_pdf_no_company(self):
        engine = ReportingEngine()
        result = engine.budget_vs_actual_to_pdf(
            company=None,
            year=2026,
            month=None,
            items=[],
            totals={k: 0.0 for k in SAMPLE_TOTALS},
        )
        self.assertIsInstance(result, bytes)
        self.assertTrue(result.startswith(b"%PDF"))


class ReportingEngineApiTests(unittest.TestCase):
    """Integration tests against the live FastAPI app."""

    def setUp(self):
        import os
        import tempfile
        from pathlib import Path
        from fastapi.testclient import TestClient
        from app import create_app

        self._temp_dir = tempfile.TemporaryDirectory()
        db_path = Path(self._temp_dir.name) / "report_test.db"

        self._original_env = {k: os.getenv(k) for k in [
            "AQ_DATABASE_PATH", "AQ_AUTH_USERS", "AQ_ENABLE_DEMO_USERS",
            "AQ_JWT_SECRET", "AQ_ENV", "AQ_MARKET_OFFLINE",
            "AQ_MACRO_OFFLINE", "AQ_WEB_OFFLINE",
        ]}
        os.environ["AQ_DATABASE_PATH"] = str(db_path)
        os.environ["AQ_AUTH_USERS"] = "admin:admin12345:admin"
        os.environ["AQ_ENABLE_DEMO_USERS"] = "false"
        os.environ["AQ_JWT_SECRET"] = "change-this-secret"
        os.environ["AQ_ENV"] = "development"
        os.environ["AQ_MARKET_OFFLINE"] = "true"
        os.environ["AQ_MACRO_OFFLINE"] = "true"
        os.environ["AQ_WEB_OFFLINE"] = "true"

        self.client = TestClient(create_app())
        resp = self.client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": "admin12345"},
        )
        self.token = resp.json()["access_token"]
        self.headers = {"Authorization": f"Bearer {self.token}"}

    def tearDown(self):
        import os
        self.client.close()
        for key, value in self._original_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        self._temp_dir.cleanup()

    def _get(self, path: str, **params):
        return self.client.get(path, params=params, headers=self.headers)

    def test_ledger_xlsx_requires_auth(self):
        resp = self.client.get("/api/v1/reports/finance/ledger.xlsx")
        self.assertEqual(resp.status_code, 401)

    @unittest.skipUnless(_OPENPYXL_OK, "openpyxl unavailable")
    def test_ledger_xlsx_returns_xlsx_content_type(self):
        resp = self._get("/api/v1/reports/finance/ledger.xlsx")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "spreadsheetml",
            resp.headers.get("content-type", ""),
        )

    @unittest.skipUnless(_OPENPYXL_OK, "openpyxl unavailable")
    def test_ledger_xlsx_has_signature_header(self):
        resp = self._get("/api/v1/reports/finance/ledger.xlsx")
        self.assertEqual(resp.status_code, 200)
        sig = resp.headers.get("x-export-signature", "")
        self.assertTrue(sig.startswith("hmac-sha256="), f"Bad sig: {sig!r}")

    @unittest.skipUnless(_OPENPYXL_OK, "openpyxl unavailable")
    def test_ledger_xlsx_signature_verifiable(self):
        resp = self._get("/api/v1/reports/finance/ledger.xlsx")
        content = resp.content
        sig = resp.headers["x-export-signature"]
        engine = ReportingEngine()
        # Default dev secret
        expected = engine.sign(content, "change-this-secret")
        self.assertEqual(sig, expected)

    def test_ledger_pdf_returns_pdf_content_type(self):
        resp = self._get("/api/v1/reports/finance/ledger.pdf")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.headers.get("content-type"), "application/pdf")

    def test_ledger_pdf_has_signature_header(self):
        resp = self._get("/api/v1/reports/finance/ledger.pdf")
        self.assertEqual(resp.status_code, 200)
        sig = resp.headers.get("x-export-signature", "")
        self.assertTrue(sig.startswith("hmac-sha256="))

    def test_budget_vs_actual_xlsx_requires_year(self):
        resp = self.client.get(
            "/api/v1/reports/finance/budget-vs-actual.xlsx",
            headers=self.headers,
        )
        self.assertEqual(resp.status_code, 422)

    @unittest.skipUnless(_OPENPYXL_OK, "openpyxl unavailable")
    def test_budget_vs_actual_xlsx_returns_xlsx(self):
        resp = self._get("/api/v1/reports/finance/budget-vs-actual.xlsx", year=2026)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp.headers.get("content-type", ""))

    def test_budget_vs_actual_pdf_returns_pdf(self):
        resp = self._get("/api/v1/reports/finance/budget-vs-actual.pdf", year=2026)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.headers.get("content-type"), "application/pdf")

    @unittest.skipUnless(_OPENPYXL_OK, "openpyxl unavailable")
    def test_content_disposition_xlsx(self):
        resp = self._get("/api/v1/reports/finance/ledger.xlsx")
        disposition = resp.headers.get("content-disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn(".xlsx", disposition)

    def test_content_disposition_pdf(self):
        resp = self._get("/api/v1/reports/finance/ledger.pdf")
        disposition = resp.headers.get("content-disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn(".pdf", disposition)


if __name__ == "__main__":
    unittest.main()
